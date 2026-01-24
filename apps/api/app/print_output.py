from __future__ import annotations

import csv
import io
import json
import logging
import os
import tempfile
from dataclasses import dataclass
from typing import Any, Optional

# Use defusedxml to prevent XXE attacks
import defusedxml.ElementTree as ET

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import Response
from PIL import Image, ImageDraw, ImageFont

from app.afp_document_generator import generate_afp_document
from app.afp_cleaner import clean_afp
from app.security import validate_file_size, validate_file_content, MAX_UPLOAD_SIZE
from app.xml_streaming_parser import (
    parse_large_xml_to_records,
    parse_large_xml_to_csv,
    get_xml_file_info,
    stream_xml_records,
)
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/print-output", tags=["print-output"])

DPI = 240
PAGE_WIDTH = int(8.5 * DPI)
PAGE_HEIGHT = int(11 * DPI)


@dataclass
class RenderConfig:
    margin_left: int = int(0.7 * DPI)
    margin_top: int = int(0.6 * DPI)
    mailing_offset_x: int = int(0.3 * DPI)
    mailing_offset_y: int = int(1.8 * DPI)
    body_start_y: int = int(3.1 * DPI)
    line_height: int = int(0.22 * DPI)


def _html_to_text(html: str) -> str:
    text = html.replace("</p>", "\n").replace("<br>", "\n").replace("<br/>", "\n")
    text = text.replace("<div>", "\n").replace("</div>", "\n")
    text = "".join(ch for ch in text if ch != "\r")
    result = []
    in_tag = False
    for ch in text:
        if ch == "<":
            in_tag = True
            continue
        if ch == ">":
            in_tag = False
            continue
        if not in_tag:
            result.append(ch)
    return "".join(result)


def _wrap_text(text: str, draw: ImageDraw.ImageDraw, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current: list[str] = []
    for word in words:
        candidate = " ".join(current + [word])
        width = draw.textlength(candidate, font=font)
        if width <= max_width or not current:
            current.append(word)
        else:
            lines.append(" ".join(current))
            current = [word]
    if current:
        lines.append(" ".join(current))
    return lines


def _replace_placeholders(text: str, row: dict[str, str], placeholder_map: dict[str, str]) -> str:
    output = text
    for placeholder, column in placeholder_map.items():
        key = placeholder.strip("[]")
        column_key = column or key
        value = row.get(column_key, "")
        output = output.replace(placeholder, value)
    return output


def _render_letter(
    body_html: str,
    block_texts: list[str],
    mailing_lines: list[str],
    return_lines: list[str],
) -> bytes:
    image = Image.new("L", (PAGE_WIDTH, PAGE_HEIGHT), color=255)
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    cfg = RenderConfig()

    x = cfg.margin_left
    y = cfg.margin_top
    for line in return_lines:
        if line:
            draw.text((x, y), line, fill=0, font=font)
            y += cfg.line_height

    mx = cfg.margin_left + cfg.mailing_offset_x
    my = cfg.margin_top + cfg.mailing_offset_y
    for line in mailing_lines:
        if line:
            draw.text((mx, my), line, fill=0, font=font)
            my += cfg.line_height

    body_text = _html_to_text(body_html).strip()
    if block_texts:
        body_text = "\n".join([body_text] + [text for text in block_texts if text.strip()])
    body_lines: list[str] = []
    for paragraph in body_text.splitlines():
        if not paragraph.strip():
            body_lines.append("")
            continue
        body_lines.extend(
            _wrap_text(
                paragraph,
                draw,
                font,
                PAGE_WIDTH - cfg.margin_left * 2,
            )
        )
    by = cfg.body_start_y
    for line in body_lines:
        if by > PAGE_HEIGHT - cfg.margin_left:
            break
        draw.text((cfg.margin_left, by), line, fill=0, font=font)
        by += cfg.line_height

    buffer = io.BytesIO()
    image.save(buffer, format="PNG", dpi=(DPI, DPI))
    return buffer.getvalue()


def _csv_from_rows(rows: list[list[Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in rows:
        writer.writerow(["" if value is None else str(value) for value in row])
    return buffer.getvalue()


def _strip_namespace(tag: str) -> str:
    """
    Strip XML namespace prefix from a tag.

    Example: "{http://www.example.com}Customer" -> "Customer"
    """
    if tag.startswith("{"):
        return tag.split("}", 1)[1] if "}" in tag else tag
    return tag


def _flatten_xml_element(element: ET.Element, prefix: str = "") -> dict[str, str]:
    """
    Flatten an XML element into a dictionary with dot-notation keys.

    Example:
        <Address><Line1>123 Main</Line1><City>NYC</City></Address>
        becomes: {"Address.Line1": "123 Main", "Address.City": "NYC"}

    Namespace prefixes are stripped from tag names.
    """
    result: dict[str, str] = {}

    # Strip namespace from tag
    tag_name = _strip_namespace(element.tag)

    # Build the current key
    current_key = f"{prefix}.{tag_name}" if prefix else tag_name

    # Get element text (strip whitespace)
    text = (element.text or "").strip()

    # Check if this element has children
    children = list(element)

    if children:
        # Has children - recurse into them
        for child in children:
            child_data = _flatten_xml_element(child, current_key)
            result.update(child_data)
    else:
        # Leaf node - add the text value
        if text:
            result[current_key] = text
        else:
            result[current_key] = ""

    # Also extract attributes as separate columns
    for attr_name, attr_value in element.attrib.items():
        attr_key = f"{current_key}@{attr_name}"
        result[attr_key] = attr_value

    return result


def _find_all_repeating_elements(element: ET.Element, depth: int = 0, max_depth: int = 3) -> list[tuple[str, list[ET.Element], int]]:
    """
    Recursively find all repeating elements in an XML structure.

    Returns a list of (tag_name, elements, count) tuples for each repeating element found.
    """
    results = []

    if depth > max_depth:
        return results

    # Count children by tag (stripped of namespace)
    tag_counts: dict[str, list[ET.Element]] = {}
    for child in element:
        tag = _strip_namespace(child.tag)
        if tag not in tag_counts:
            tag_counts[tag] = []
        tag_counts[tag].append(child)

    # Find tags with multiple occurrences
    for tag, elements in tag_counts.items():
        if len(elements) > 1:
            results.append((tag, elements, len(elements)))

    # Recursively check children
    for child in element:
        child_results = _find_all_repeating_elements(child, depth + 1, max_depth)
        results.extend(child_results)

    return results


def _detect_repeating_element(root: ET.Element) -> tuple[str, list[ET.Element]]:
    """
    Auto-detect the repeating element (records) in an XML structure.

    Looks for repeating elements at any level up to 3 levels deep.
    If multiple repeating elements are found, picks the one with the most occurrences.
    Returns the tag name and list of record elements.
    """
    # Find all repeating elements
    all_repeating = _find_all_repeating_elements(root)

    if not all_repeating:
        # Fallback: treat all direct children as records
        children = list(root)
        if children:
            return _strip_namespace(children[0].tag), children
        return "", []

    # Sort by count (most occurrences first)
    all_repeating.sort(key=lambda x: x[2], reverse=True)

    # Return the repeating element with most occurrences
    tag, elements, count = all_repeating[0]
    return tag, elements


def _parse_xml_to_records(xml_content: str) -> tuple[list[str], list[dict[str, str]]]:
    """
    Parse XML content and extract columns and records.

    Returns:
        tuple of (columns list, records list)
        - columns: list of all unique column names (dot-notation for nested)
        - records: list of dictionaries, each representing one record
    """
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        raise ValueError(f"Invalid XML: {e}")

    # Detect the repeating element
    record_tag, record_elements = _detect_repeating_element(root)

    if not record_elements:
        raise ValueError("Could not find repeating records in XML")

    # Extract all records
    records: list[dict[str, str]] = []
    all_columns: set[str] = set()

    for record_elem in record_elements:
        # Flatten this record (don't include the record tag itself in the path)
        record_data: dict[str, str] = {}

        for child in record_elem:
            child_data = _flatten_xml_element(child)
            record_data.update(child_data)

        # Also get attributes of the record element itself
        for attr_name, attr_value in record_elem.attrib.items():
            record_data[f"@{attr_name}"] = attr_value

        records.append(record_data)
        all_columns.update(record_data.keys())

    # Sort columns for consistent ordering
    columns = sorted(list(all_columns))

    # Ensure all records have all columns (fill missing with empty string)
    for record in records:
        for col in columns:
            if col not in record:
                record[col] = ""

    return columns, records


def _xml_records_to_csv(columns: list[str], records: list[dict[str, str]]) -> str:
    """
    Convert XML records to CSV format.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Write header
    writer.writerow(columns)

    # Write data rows
    for record in records:
        row = [record.get(col, "") for col in columns]
        writer.writerow(row)

    return buffer.getvalue()


def _flatten_json_object(obj: dict[str, Any], prefix: str = "") -> dict[str, str]:
    """
    Flatten a nested JSON object into a flat dictionary.
    Nested keys are joined with underscore: {"address": {"city": "NYC"}} -> {"address_city": "NYC"}
    """
    result: dict[str, str] = {}
    for key, value in obj.items():
        new_key = f"{prefix}_{key}" if prefix else key
        if isinstance(value, dict):
            result.update(_flatten_json_object(value, new_key))
        elif isinstance(value, list):
            # Convert lists to comma-separated strings
            result[new_key] = ", ".join(str(v) for v in value)
        elif value is None:
            result[new_key] = ""
        else:
            result[new_key] = str(value)
    return result


def _parse_json_to_records(json_text: str) -> tuple[list[str], list[dict[str, str]]]:
    """
    Parse JSON text and extract records.

    Supports:
    - Array of objects: [{...}, {...}]
    - Object with data array: {"data": [{...}]} or {"records": [{...}]}
    - Single object (treated as one record)
    """
    data = json.loads(json_text)
    records: list[dict[str, Any]] = []

    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # Look for common data array keys
        data_keys = ["data", "records", "items", "rows", "results"]
        for key in data_keys:
            if key in data and isinstance(data[key], list):
                records = data[key]
                break

        # If no data array found, look for any array value
        if not records:
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    records = value
                    break

        # If still no records, treat the object itself as a single record
        if not records and data:
            records = [data]

    if not records:
        raise ValueError("No records found in JSON file")

    # Flatten all records
    flattened: list[dict[str, str]] = []
    all_columns: set[str] = set()

    for record in records:
        if isinstance(record, dict):
            flat = _flatten_json_object(record)
            flattened.append(flat)
            all_columns.update(flat.keys())

    # Sort columns for consistent ordering
    columns = sorted(all_columns)

    return columns, flattened


def _json_records_to_csv(columns: list[str], records: list[dict[str, str]]) -> str:
    """
    Convert JSON records to CSV format.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Write header
    writer.writerow(columns)

    # Write data rows
    for record in records:
        row = [record.get(col, "") for col in columns]
        writer.writerow(row)

    return buffer.getvalue()


@router.post("/columns")
async def parse_columns(file: UploadFile = File(...)) -> dict[str, Any]:
    """
    Parse uploaded file and extract columns.

    Supports:
    - CSV files (.csv)
    - Excel files (.xlsx)
    - XML files (.xml) - auto-detects repeating elements as records
    - JSON files (.json) - supports arrays, nested objects, and common data wrappers

    Returns:
        - columns: list of column/field names
        - csv: data in CSV format (for uniform processing)
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing file name")

    # Validate file size before reading
    if file.size is not None:
        validate_file_size(file.size)

    filename = file.filename.lower()

    # Read file with size limit
    data = await file.read()
    if len(data) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {MAX_UPLOAD_SIZE // (1024 * 1024)} MB",
        )

    # Determine expected file type and validate content
    if filename.endswith(".csv"):
        expected_type = "csv"
    elif filename.endswith(".xlsx"):
        expected_type = "xlsx"
    elif filename.endswith(".xml"):
        expected_type = "xml"
    elif filename.endswith(".json"):
        expected_type = "json"
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Supported: .csv, .xlsx, .xml, .json",
        )

    # Validate file content matches extension
    try:
        validate_file_content(data, expected_type)
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"File content validation failed: {e}")
        raise HTTPException(
            status_code=400,
            detail="File content does not match the file extension",
        )

    # Parse based on file type
    if expected_type == "csv":
        text = data.decode("utf-8", errors="ignore")
        header = text.splitlines()[0] if text.splitlines() else ""
        columns = [value.strip() for value in header.split(",") if value.strip()]
        return {"columns": columns, "csv": text}

    if expected_type == "xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        if not rows:
            raise HTTPException(status_code=400, detail="Spreadsheet has no rows")
        columns = [str(value).strip() for value in rows[0] if value is not None]
        return {"columns": columns, "csv": _csv_from_rows(rows)}

    if expected_type == "xml":
        try:
            xml_text = data.decode("utf-8", errors="ignore")
            columns, records = _parse_xml_to_records(xml_text)
            if not records:
                raise HTTPException(status_code=400, detail="XML file has no records")
            csv_text = _xml_records_to_csv(columns, records)
            return {"columns": columns, "csv": csv_text}
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    if expected_type == "json":
        try:
            json_text = data.decode("utf-8", errors="ignore")
            columns, records = _parse_json_to_records(json_text)
            if not records:
                raise HTTPException(status_code=400, detail="JSON file has no records")
            csv_text = _json_records_to_csv(columns, records)
            return {"columns": columns, "csv": csv_text}
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # This should never be reached due to earlier validation
    raise HTTPException(status_code=400, detail="Unsupported file type")


@router.post("/afp")
def generate_afp(payload: dict[str, Any]) -> Response:
    """
    Generate a complete IBM AFP document with embedded TLE index records.

    The AFP document contains:
    - Document structure (BDT/EDT)
    - Page structure with page descriptors
    - TLE (Tag Logical Element) records for each page containing:
      - mailing_name, mailing_addr1, mailing_addr2, mailing_addr3
      - return_addr1, return_addr2, return_addr3
    - Inline page segments with IOCA image data

    Compatible with mainframe processing, reblocking, and AFP viewers.
    """
    try:
        spreadsheet_csv = payload.get("spreadsheet_csv", "")
        if not spreadsheet_csv.strip():
            raise HTTPException(status_code=400, detail="Spreadsheet data missing")
        reader = csv.DictReader(io.StringIO(spreadsheet_csv))
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=400, detail="Spreadsheet has no rows")

        body_html = payload.get("template_html", "")
        block_texts = payload.get("block_texts", [])
        placeholder_map = payload.get("placeholder_map", {})
        mailing_map = payload.get("mailing_map", {})
        return_lines = payload.get("return_address", ["", "", ""])

        # Dynamic asset configuration
        dynamic_return = payload.get("dynamic_return")

        # Helper to get return lines for a row
        def get_return_lines_for_row(row: dict[str, str]) -> list[str]:
            if dynamic_return:
                column = dynamic_return.get("column", "")
                value_map = dynamic_return.get("map", {})
                default = dynamic_return.get("default", ["", "", ""])
                if column and value_map:
                    value = row.get(column, "")
                    if value in value_map:
                        return value_map[value]
                return default if default else ["", "", ""]
            return return_lines

        # Build pages with image data and TLE information
        pages: list[dict[str, Any]] = []
        for index, row in enumerate(rows, start=1):
            merged_body = _replace_placeholders(body_html, row, placeholder_map)
            merged_blocks = [
                _replace_placeholders(text, row, placeholder_map) for text in block_texts
            ]

            mailing_lines = [
                row.get(mailing_map.get("mailing_name", ""), ""),
                row.get(mailing_map.get("mailing_addr1", ""), ""),
                row.get(mailing_map.get("mailing_addr2", ""), ""),
                row.get(mailing_map.get("mailing_addr3", ""), ""),
            ]

            # Get return lines for this row (static or dynamic)
            row_return_lines = get_return_lines_for_row(row)

            image_bytes = _render_letter(
                merged_body,
                merged_blocks,
                mailing_lines,
                row_return_lines,
            )

            # Convert PNG to grayscale image data
            image = Image.open(io.BytesIO(image_bytes))
            if image.mode != "L":
                image = image.convert("L")
            image_data = image.tobytes()
            width, height = image.size

            # Build page with TLE data
            pages.append({
                'image_data': image_data,
                'width': width,
                'height': height,
                'tle_data': {
                    'mailing_name': mailing_lines[0],
                    'mailing_addr1': mailing_lines[1],
                    'mailing_addr2': mailing_lines[2],
                    'mailing_addr3': mailing_lines[3],
                    'return_addr1': row_return_lines[0] if len(row_return_lines) > 0 else "",
                    'return_addr2': row_return_lines[1] if len(row_return_lines) > 1 else "",
                    'return_addr3': row_return_lines[2] if len(row_return_lines) > 2 else "",
                }
            })

        # Generate complete AFP document with embedded TLE records
        afp_document = generate_afp_document(
            pages=pages,
            document_name="MAILOUT",
            resolution=DPI,
            page_width=PAGE_WIDTH,
            page_height=PAGE_HEIGHT
        )

        return Response(
            content=afp_document,
            media_type="application/octet-stream",
            headers={"Content-Disposition": 'attachment; filename="print_output.afp"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/pdf")
def generate_pdf(payload: dict[str, Any]) -> Response:
    """
    Generate a PDF document with rendered letters.

    Each row in the spreadsheet becomes a page in the PDF.
    """
    try:
        spreadsheet_csv = payload.get("spreadsheet_csv", "")
        if not spreadsheet_csv.strip():
            raise HTTPException(status_code=400, detail="Spreadsheet data missing")
        reader = csv.DictReader(io.StringIO(spreadsheet_csv))
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=400, detail="Spreadsheet has no rows")

        body_html = payload.get("template_html", "")
        block_texts = payload.get("block_texts", [])
        placeholder_map = payload.get("placeholder_map", {})
        mailing_map = payload.get("mailing_map", {})
        return_lines = payload.get("return_address", ["", "", ""])

        # Dynamic asset configuration
        dynamic_return = payload.get("dynamic_return")

        # Helper to get return lines for a row
        def get_return_lines_for_row(row: dict[str, str]) -> list[str]:
            if dynamic_return:
                column = dynamic_return.get("column", "")
                value_map = dynamic_return.get("map", {})
                default = dynamic_return.get("default", ["", "", ""])
                if column and value_map:
                    value = row.get(column, "")
                    if value in value_map:
                        return value_map[value]
                return default if default else ["", "", ""]
            return return_lines

        # Render each row as a page image
        images: list[Image.Image] = []
        for index, row in enumerate(rows, start=1):
            merged_body = _replace_placeholders(body_html, row, placeholder_map)
            merged_blocks = [
                _replace_placeholders(text, row, placeholder_map) for text in block_texts
            ]

            mailing_lines = [
                row.get(mailing_map.get("mailing_name", ""), ""),
                row.get(mailing_map.get("mailing_addr1", ""), ""),
                row.get(mailing_map.get("mailing_addr2", ""), ""),
                row.get(mailing_map.get("mailing_addr3", ""), ""),
            ]

            # Get return lines for this row (static or dynamic)
            row_return_lines = get_return_lines_for_row(row)

            image_bytes = _render_letter(
                merged_body,
                merged_blocks,
                mailing_lines,
                row_return_lines,
            )

            # Open rendered image and convert to RGB for PDF
            image = Image.open(io.BytesIO(image_bytes))
            if image.mode != "RGB":
                image = image.convert("RGB")
            images.append(image)

        # Save all images as a multi-page PDF
        pdf_buffer = io.BytesIO()
        if images:
            first_image = images[0]
            additional_images = images[1:] if len(images) > 1 else []
            first_image.save(
                pdf_buffer,
                format="PDF",
                save_all=True,
                append_images=additional_images,
                resolution=DPI,
            )

        pdf_buffer.seek(0)
        return Response(
            content=pdf_buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="print_output.pdf"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/afp-clean")
async def clean_afp_file(file: UploadFile = File(...)) -> Response:
    """
    Remove index/grouping structured fields from an AFP file.

    This endpoint removes Named Page Group and Index Element structured fields
    that are added by mainframe processing (e.g., StreamWeaver) and may not be
    recognized by downstream systems like Bluecrest Output Manager.

    Structured fields removed:
    - BNG (D3 A8 5F) - Begin Named Page Group
    - ENG (D3 A9 5F) - End Named Page Group
    - BIE (D3 A8 FB) - Begin Index Element
    - EIE (D3 A9 FB) - End Index Element
    - IEL (D3 AF 5F) - Index Element Link

    Only removes these fields when they contain S-number patterns (S0000001, etc.).
    """
    try:
        # Read the uploaded file
        content = await file.read()

        # Validate file size (max 50MB for AFP files)
        max_afp_size = 50 * 1024 * 1024  # 50MB
        if len(content) > max_afp_size:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum size is {max_afp_size // (1024*1024)}MB"
            )

        # Validate it's an AFP file (starts with 0x5A carriage control)
        if len(content) < 6 or content[0] != 0x5A:
            raise HTTPException(
                status_code=400,
                detail="Invalid AFP file format. File must start with AFP carriage control (0x5A)"
            )

        # Clean the AFP file
        cleaned_data, stats = clean_afp(content)

        # Generate filename
        original_name = file.filename or "output"
        if original_name.lower().endswith('.afp'):
            clean_name = original_name[:-4] + ".cleaned.afp"
        else:
            clean_name = original_name + ".cleaned.afp"

        logger.info(
            f"AFP cleaned: removed {stats['total_removed']} index fields, "
            f"kept {stats['total_kept']} structured fields"
        )

        return Response(
            content=cleaned_data,
            media_type="application/octet-stream",
            headers={
                "Content-Disposition": f'attachment; filename="{clean_name}"',
                "X-AFP-Removed-BNG": str(stats.get("BNG (Begin Named Page Group)", 0)),
                "X-AFP-Removed-ENG": str(stats.get("ENG (End Named Page Group)", 0)),
                "X-AFP-Removed-BIE": str(stats.get("BIE (Begin Index Element)", 0)),
                "X-AFP-Removed-IEL": str(stats.get("IEL (Index Element Link)", 0)),
                "X-AFP-Total-Removed": str(stats["total_removed"]),
                "X-AFP-Total-Kept": str(stats["total_kept"]),
            },
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"AFP cleaning failed: {exc}")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


# Maximum size for large XML files (2GB)
MAX_LARGE_XML_SIZE = 2 * 1024 * 1024 * 1024  # 2GB


@router.post("/xml-large/info")
async def get_large_xml_info(file: UploadFile = File(...)) -> dict[str, Any]:
    """
    Get information about a large XML file without fully parsing it.

    This endpoint analyzes the XML structure and estimates the record count
    without loading the entire file into memory.

    Returns:
        - file_size_bytes: File size in bytes
        - file_size_mb: File size in megabytes
        - detected_record_tag: The detected repeating element tag
        - estimated_records: Estimated number of records
    """
    if not file.filename or not file.filename.lower().endswith('.xml'):
        raise HTTPException(status_code=400, detail="File must be an XML file (.xml)")

    # Save to temp file for streaming analysis
    temp_file = None
    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xml')
        temp_path = temp_file.name

        # Stream file to disk in chunks
        total_size = 0
        chunk_size = 64 * 1024  # 64KB chunks

        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_LARGE_XML_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large. Maximum size is {MAX_LARGE_XML_SIZE // (1024*1024*1024)}GB"
                )
            temp_file.write(chunk)

        temp_file.close()

        # Get file info using streaming parser
        info = get_xml_file_info(temp_path)
        info["filename"] = file.filename

        return info

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"XML info extraction failed: {exc}")
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        if temp_file:
            try:
                os.unlink(temp_file.name)
            except Exception:
                pass


@router.post("/xml-large/parse")
async def parse_large_xml(
    file: UploadFile = File(...),
    max_records: Optional[int] = 10000,
    record_tag: Optional[str] = None,
) -> dict[str, Any]:
    """
    Parse a large XML file using streaming (memory-efficient).

    This endpoint can handle XML files up to 2GB by using iterative parsing
    that processes records one at a time without loading the entire file.

    Args:
        file: The XML file to parse
        max_records: Maximum records to return (default 10000, max 50000)
        record_tag: Override the detected record element tag (optional)

    Returns:
        - columns: list of column names
        - csv: data in CSV format
        - records_processed: number of records processed
        - truncated: whether the result was truncated due to max_records
    """
    if not file.filename or not file.filename.lower().endswith('.xml'):
        raise HTTPException(status_code=400, detail="File must be an XML file (.xml)")

    # Limit max_records to prevent memory issues
    if max_records is None or max_records > 50000:
        max_records = 50000

    temp_file = None
    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xml')
        temp_path = temp_file.name

        # Stream file to disk
        total_size = 0
        chunk_size = 64 * 1024

        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_LARGE_XML_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large. Maximum size is {MAX_LARGE_XML_SIZE // (1024*1024*1024)}GB"
                )
            temp_file.write(chunk)

        temp_file.close()

        logger.info(f"Parsing large XML file: {file.filename} ({total_size / (1024*1024):.1f} MB)")

        # Parse using streaming parser
        columns, records = parse_large_xml_to_records(
            temp_path,
            record_tag=record_tag,
            max_records=max_records
        )

        # Convert to CSV format
        csv_text = _xml_records_to_csv(columns, records)

        truncated = len(records) >= max_records

        logger.info(f"Parsed {len(records)} records with {len(columns)} columns (truncated={truncated})")

        return {
            "columns": columns,
            "csv": csv_text,
            "records_processed": len(records),
            "truncated": truncated,
            "file_size_mb": round(total_size / (1024 * 1024), 2),
        }

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as exc:
        logger.error(f"Large XML parsing failed: {exc}")
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        if temp_file:
            try:
                os.unlink(temp_file.name)
            except Exception:
                pass


@router.post("/xml-large/to-csv")
async def convert_large_xml_to_csv(
    file: UploadFile = File(...),
    record_tag: Optional[str] = None,
) -> Response:
    """
    Convert a large XML file to CSV format (streaming).

    This endpoint handles XML files up to 2GB by streaming the data
    and returning a downloadable CSV file.

    Args:
        file: The XML file to convert
        record_tag: Override the detected record element tag (optional)

    Returns:
        CSV file download
    """
    if not file.filename or not file.filename.lower().endswith('.xml'):
        raise HTTPException(status_code=400, detail="File must be an XML file (.xml)")

    temp_xml = None
    temp_csv = None
    try:
        # Save XML to temp file
        temp_xml = tempfile.NamedTemporaryFile(delete=False, suffix='.xml')
        xml_path = temp_xml.name

        total_size = 0
        chunk_size = 64 * 1024

        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_LARGE_XML_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large. Maximum size is {MAX_LARGE_XML_SIZE // (1024*1024*1024)}GB"
                )
            temp_xml.write(chunk)

        temp_xml.close()

        logger.info(f"Converting large XML to CSV: {file.filename} ({total_size / (1024*1024):.1f} MB)")

        # Convert to CSV
        temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        csv_path = temp_csv.name
        temp_csv.close()

        csv_path, record_count, columns = parse_large_xml_to_csv(
            xml_path,
            output_path=csv_path,
            record_tag=record_tag
        )

        logger.info(f"Converted {record_count} records with {len(columns)} columns")

        # Read CSV and return
        with open(csv_path, 'rb') as f:
            csv_content = f.read()

        # Generate output filename
        output_name = file.filename[:-4] + ".csv" if file.filename else "output.csv"

        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{output_name}"',
                "X-Records-Count": str(record_count),
                "X-Columns-Count": str(len(columns)),
            },
        )

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as exc:
        logger.error(f"Large XML to CSV conversion failed: {exc}")
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        for temp in [temp_xml, temp_csv]:
            if temp:
                try:
                    os.unlink(temp.name)
                except Exception:
                    pass
